import os
os.environ['NLTK_DATA'] = '/opt/render/nltk_data'

import nltk
nltk.download('stopwords', quiet=True)
nltk.download('punkt', quiet=True)
nltk.download('wordnet', quiet=True)
nltk.download('omw-1.4', quiet=True)

import io
import requests
from flask import Flask, request, redirect, session
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)

AI_SERVICES = {
    'OpenAI': {
        'guide': [
            '1. Acesse https://platform.openai.com/ e faça login',
            '2. Clique em "API Keys" no menu lateral',
            '3. Clique em "Create new secret key"',
            '4. Nomeie a chave (ex: MeuSistema) e clique em "Create"',
            '5. Copie a chave gerada e cole abaixo'
        ],
        'api_url': 'https://api.openai.com/v1/chat/completions'
    },
    'HuggingFace': {
        'guide': [
            '1. Acesse https://huggingface.co/ e faça login',
            '2. Clique em seu ícone de perfil > Settings > Access Tokens',
            '3. Gere um token com permissão "Read"',
            '4. Escolha um modelo na lista abaixo',
            '5. Cole o token e selecione o modelo desejado'
        ],
        'api_url': 'https://api-inference.huggingface.co/models/'
    },
    'Cohere': {
        'guide': [
            '1. Acesse https://dashboard.cohere.ai/ e faça login',
            '2. Navegue até a seção "API Keys"',
            '3. Clique em "Generate API Key"',
            '4. Nomeie a chave (ex: MeuSistema) e confirme',
            '5. Copie a chave gerada e cole abaixo'
        ],
        'api_url': 'https://api.cohere.ai/v1/generate'
    }
}

MODELOS_HF = {
    'português': [
        ('unicamp-dl/ptt5-base-portuguese-vocab', 'PTT5 (Sumarização em Português)'),
        ('pierreguillou/bert-base-cased-squad-v1.1-portuguese', 'BERT PT (Q&A)'),
        ('neuralmind/bert-base-portuguese-cased', 'BERT PT (Geral)')
    ],
    'multilíngue': [
        ('google/mt5-base', 'mT5 (Multilíngue)'),
        ('facebook/bart-large-cnn', 'BART (Inglês)'),
        ('google/flan-t5-large', 'FLAN-T5 (Inglês)')
    ]
}

HTML_BASE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Sistema de Resumo de Documentos</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .container {{ max-width: 800px; margin: 50px auto; }}
        .card {{ margin-top: 20px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); }}
        .guide-step {{ margin: 15px 0; padding: 10px; background: #f8f9fa; border-radius: 5px; }}
        .btn-custom {{ margin: 5px; padding: 15px 30px; font-size: 1.1em; }}
        pre {{ white-space: pre-wrap; word-wrap: break-word; background: #f8f9fa; padding: 15px; border-radius: 5px; }}
        .model-group {{ margin-bottom: 20px; }}
        .model-option {{ padding: 10px; border: 1px solid #dee2e6; border-radius: 5px; margin-bottom: 10px; }}
    </style>
</head>
<body>
    <div class="container">
        <h2 class="text-center mb-4">📄 Sistema de Resumo de Documentos</h2>
        <div class="card p-4">
            <a href="/" class="btn btn-secondary mb-3">🏠 Voltar</a>
            {}
        </div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
'''

def extract_text(file):
    filename = file.filename
    content = file.read()
    
    if filename.endswith('.pdf'):
        pdf = PdfReader(io.BytesIO(content))
        return '\n'.join([page.extract_text() for page in pdf.pages])
    
    elif filename.endswith('.docx'):
        doc = Document(io.BytesIO(content))
        return '\n'.join([para.text for para in doc.paragraphs])
    
    elif filename.endswith(('.xlsx', '.xls')):
        wb = load_workbook(io.BytesIO(content))
        text = []
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                text.append(' '.join(map(str, row)))
        return '\n'.join(text)
    
    return ''

def generate_summary(text, filename):
    service = session.get('ai_service')
    api_key = session.get('api_key')
    
    prompt = (
        f"Resuma este documento de forma clara e detalhada em português brasileiro.\n"
        f"Documento: {filename}\n"
        f"Conteúdo:\n{text}\n\n"
        "Instruções:\n"
        "- Mantenha termos técnicos importantes\n"
        "- Use estrutura com tópicos\n"
        "- Limite de 300 palavras\n"
    )
    
    try:
        if service == 'OpenAI':
            headers = {'Authorization': f'Bearer {api_key}'}
            data = {
                "model": "gpt-3.5-turbo",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7
            }
            response = requests.post(
                AI_SERVICES[service]['api_url'],
                json=data,
                headers=headers,
                timeout=30
            )
            return response.json()['choices'][0]['message']['content']
        
        elif service == 'HuggingFace':
            headers = {'Authorization': f'Bearer {api_key}'}
            modelo = session.get('hf_model', 'unicamp-dl/ptt5-base-portuguese-vocab')
            
            data = {
                "inputs": prompt,
                "parameters": {
                    "max_length": 512,
                    "min_length": 150,
                    "temperature": 0.9,
                    "repetition_penalty": 2.5,
                    "num_beams": 4
                }
            }
            
            response = requests.post(
                f"{AI_SERVICES[service]['api_url']}{modelo}",
                json=data,
                headers=headers,
                timeout=60
            )
            
            if response.status_code == 503:
                return "🔃 Modelo está carregando. Recarregue a página em 30 segundos."
                
            if response.status_code != 200:
                return f"Erro {response.status_code}: {response.text[:300]}"
                
            return response.json()[0]['generated_text']
        
        elif service == 'Cohere':
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            }
            data = {
                "prompt": prompt,
                "max_tokens": 1000,
                "temperature": 0.7,
                "model": "command-nightly"
            }
            response = requests.post(
                AI_SERVICES[service]['api_url'],
                json=data,
                headers=headers,
                timeout=30
            )
            return response.json()['generations'][0]['text']
    
    except requests.exceptions.RequestException as e:
        return f"⛔ Erro de conexão: {str(e)}"
    except KeyError as e:
        return f"⚠️ Erro no formato da resposta: {str(e)}"
    except Exception as e:
        return f"⚠️ Erro inesperado: {str(e)}"

@app.route('/')
def home():
    content = '''
    <div class="text-center">
        <h4 class="mb-4">Selecione uma opção:</h4>
        <a href="/settings" class="btn btn-primary btn-custom">⚙️ Configurações</a>
        <a href="/process" class="btn btn-success btn-custom">📄 Processar Documento</a>
    </div>
    '''
    return HTML_BASE.format(content)

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        session['ai_service'] = request.form.get('ai_service')
        return redirect(f'/configure/{request.form.get("ai_service")}')
    
    content = '''
    <h4 class="mb-4">🔧 Configurações do Sistema</h4>
    <form method="POST">
        <div class="mb-3">
            <label class="form-label">Selecione o serviço de IA:</label>
            <select name="ai_service" class="form-select" required>
                <option value="">-- Selecione --</option>
                <option value="OpenAI">OpenAI (GPT-3.5)</option>
                <option value="HuggingFace">HuggingFace</option>
                <option value="Cohere">Cohere</option>
            </select>
        </div>
        <button type="submit" class="btn btn-primary">Avançar</button>
    </form>
    '''
    return HTML_BASE.format(content)

@app.route('/configure/<service>', methods=['GET', 'POST'])
def configure(service):
    if request.method == 'POST':
        session['api_key'] = request.form.get('api_key')
        if service == 'HuggingFace':
            session['hf_model'] = request.form.get('hf_model')
        return redirect('/')
    
    if service == 'HuggingFace':
        content = f'''
        <h4 class="mb-4">🔧 Configurar HuggingFace</h4>
        <div class="mb-4">
            <h5>Siga estas instruções:</h5>
            <div class="guide-step">
                {'</div><div class="guide-step">'.join(AI_SERVICES[service]['guide'])}
            </div>
        </div>
        <form method="POST">
            <div class="mb-3">
                <label class="form-label">Token de Acesso:</label>
                <input type="text" name="api_key" class="form-control" required>
            </div>
            <div class="mb-3">
                <label class="form-label">Selecione o Modelo:</label>
                <select name="hf_model" class="form-select" required>
                    <optgroup label="Modelos em Português">
                        {"".join(f'<option value="{m[0]}">{m[1]}</option>' for m in MODELOS_HF["português"])}
                    </optgroup>
                    <optgroup label="Modelos Multilíngues">
                        {"".join(f'<option value="{m[0]}">{m[1]}</option>' for m in MODELOS_HF["multilíngue"])}
                    </optgroup>
                </select>
            </div>
            <button type="submit" class="btn btn-primary">Salvar</button>
        </form>
        '''
    else:
        guide = AI_SERVICES.get(service, {}).get('guide', [])
        content = f'''
        <h4 class="mb-4">🔑 Configurar {service}</h4>
        <div class="mb-4">
            <h5>Siga estas instruções:</h5>
            {'<hr>'.join(f'<div class="guide-step">{step}</div>' for step in guide)}
        </div>
        <form method="POST">
            <div class="mb-3">
                <label class="form-label">Cole sua chave API aqui:</label>
                <input type="text" name="api_key" class="form-control" required>
            </div>
            <button type="submit" class="btn btn-primary">Salvar</button>
        </form>
        '''
    
    return HTML_BASE.format(content)

@app.route('/process', methods=['GET', 'POST'])
def process():
    if 'api_key' not in session:
        return redirect('/settings')
    
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        
        text = extract_text(file)
        summary = generate_summary(text, file.filename)
        
        content = f'''
        <h4 class="mb-4">📝 Resumo Gerado</h4>
        <div class="alert alert-success">
            <h5>{file.filename}</h5>
            <pre>{summary}</pre>
        </div>
        <a href="/process" class="btn btn-primary">Nova Análise</a>
        '''
        return HTML_BASE.format(content)
    
    content = '''
    <h4 class="mb-4">📤 Processar Documento</h4>
    <form method="POST" enctype="multipart/form-data">
        <div class="mb-3">
            <label class="form-label">Selecione o arquivo (PDF, DOCX, XLSX):</label>
            <input type="file" name="file" class="form-control" accept=".pdf,.docx,.xlsx" required>
        </div>
        <button type="submit" class="btn btn-primary">Enviar e Processar</button>
    </form>
    '''
    return HTML_BASE.format(content)

if __name__ == '__main__':
    app.run(debug=True)
    app.run(debug=True)

